import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from zaevochnik_engine.formulas import get_excel_col_letter


def apply_top_header_and_protection(worksheet, start_row: int, end_row: int, column_mapping: dict,
                                    weight_column_name: str, excel_styles: dict):
    """
    Формирует и стилизует верхнюю шапку (строки 1-4) на основе скриншота с внешними границами,
    добавляет префикс формата веса, включает защиту ячеек, а также автоподбор ширины и высоты для шапки.
    """
    cfg = excel_styles["top_header"]
    bg_fill = PatternFill(start_color=cfg["bg_color"], end_color=cfg["bg_color"], fill_type="solid")

    # Стили шрифтов Cambria
    font_company = Font(name=cfg["font_name"],
                        size=cfg["company_name_font_size"] if "company_name_font_size" in cfg else cfg[
                            "company_font_size"], bold=True, color=cfg["text_color_white"])
    font_yellow = Font(name=cfg["font_name"], size=cfg["label_font_size"], bold=True, color=cfg["text_color_yellow"])
    font_white = Font(name=cfg["font_name"], size=cfg["label_font_size"], bold=True, color=cfg["text_color_white"])

    # Добавляем wrap_text=True для левого и правого выравнивания шапки, чтобы автоподбор высоты корректно работал
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Настройка стиля для внешней границы (цвет берется из конфигурации)
    border_color = cfg.get("border", "000000")  # Черный по умолчанию, если ключа нет
    thin_side = Side(border_style="thin", color=border_color)

    # 1. Снимаем старые границы, красим фон и накладываем ВНЕШНИЙ контур для блока А1:I4
    for r in range(1, 5):
        for c in range(1, 10):
            cell = worksheet.cell(row=r, column=c)
            cell.fill = bg_fill
            cell.protection = Protection(locked=True)  # По умолчанию заблокировано

            # Формируем внешнюю границу по периметру диапазона A1:I4
            cell.border = Border(
                top=thin_side if r == 1 else None,
                bottom=thin_side if r == 4 else None,
                left=thin_side if c == 1 else None,
                right=thin_side if c == 9 else None
            )

    # 2. Наполнение структуры текстом и объединение по макету скриншота
    # Строка 1: Последнее обновление
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = f"Последнее обновление заявочника: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    worksheet["A1"].font = font_yellow
    worksheet["A1"].alignment = align_left

    # Центральный блок: Мунай Пром (строки 1-4, колонки C-E)
    worksheet.merge_cells("C1:E4")
    worksheet["C1"] = cfg["company_name"]
    worksheet["C1"].font = font_company
    worksheet["C1"].alignment = align_center

    # Строка 3: АЗС № и поле ввода
    worksheet["A2"] = "АЗС №"
    worksheet["A2"].font = font_white
    worksheet["A2"].alignment = align_right

    worksheet["B2"] = "[ введите номер ]"
    worksheet["B2"].font = font_white
    worksheet["B2"].alignment = align_left
    worksheet["B2"].protection = Protection(locked=False)  # Разблокировано для ввода!

    # Строка 4: Дата заявки
    worksheet["A3"] = "Дата заявки:"
    worksheet["A3"].font = font_white
    worksheet["A3"].alignment = align_right

    worksheet["B3"] = "[ введите дату ]"
    worksheet["B3"].font = font_white
    worksheet["B3"].alignment = align_left
    worksheet["B3"].protection = Protection(locked=False)

    # Правый блок: Общий вес и его сумма (колонки H и I)
    worksheet.merge_cells("H1:H4")
    worksheet["H1"] = "Общий вес:"
    worksheet["H1"].font = font_white
    worksheet["H1"].alignment = align_center

    worksheet.merge_cells("I1:I4")
    worksheet["I1"].font = font_white
    worksheet["I1"].alignment = align_center

    # Записываем формулу суммы веса динамически на основе буквы колонки
    f_total_weight_letter = get_excel_col_letter(worksheet, start_row, weight_column_name)
    worksheet["I1"] = f"=SUM({f_total_weight_letter}{start_row + 1}:{f_total_weight_letter}{end_row})"

    # Суффикс для ячейки общей суммы в шапке
    worksheet["I1"].number_format = '#,##0.00" кг"'

    # 4. Тотальная защита листа, кроме столбца "Кол-во заказа"
    qty_letter = get_excel_col_letter(worksheet, start_row, column_mapping["qty_col"])

    # Разблокируем ячейки данных в столбце "Кол-во заказа"
    for row in range(start_row + 1, end_row + 1):
        worksheet[f"{qty_letter}{row}"].protection = Protection(locked=False)


    # ─── НАСТРОЙКА АВТОПОДБОРА ШИРИНЫ КОЛОНОК ШАПКИ (A и B) ─────────────
    for col_idx in [1, 2]:
        col_letter = get_column_letter(col_idx)
        max_header_len = 0

        for row_idx in range(1, 5):
            cell_val = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and not str(cell_val).startswith("="):
                max_header_len = max(max_header_len, len(str(cell_val)))

        header_calculated_width = min(max_header_len + 5, 20)  # Слегка увеличили лимит до 20

        current_width = worksheet.column_dimensions[col_letter].width or 10
        worksheet.column_dimensions[col_letter].width = max(current_width, header_calculated_width)

    # ─── ДИНАМИЧЕСКИЙ АВТОПОДБОР ВЫСОТЫ СТРОК ДЛЯ ШАПКИ (2-4) ───────────
    # Строку 1 исключаем (range(2, 5)), чтобы длинный текст обновления в A1:B1 не раздувал высоту
    for r in range(2, 5):
        max_row_height = 22  # Базовая комфортная высота строки для текста в один ряд

        for c in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=r, column=c)
            # Проверяем только значащие текстовые ячейки в колонках A и B
            if cell.value is not None and not str(cell.value).startswith("=") and c <= 2:
                val_str = str(cell.value)
                col_letter = get_column_letter(c)
                col_width = worksheet.column_dimensions[col_letter].width or 12

                if len(val_str) > col_width:
                    lines_count = (len(val_str) // int(col_width)) + 1
                    calculated_height = lines_count * 16
                    max_row_height = max(max_row_height, calculated_height)

        worksheet.row_dimensions[r].height = max_row_height

    # А для первой строки жестко задаем стандартную красивую высоту
    worksheet.row_dimensions[1].height = 24

    # Включаем аппаратную защиту самого листа в Excel
    worksheet.protection.password = cfg["sheet_password"]
    worksheet.protection.selectLockedCells = True
    worksheet.protection.selectUnlockedCells = False
    worksheet.protection.enable()
