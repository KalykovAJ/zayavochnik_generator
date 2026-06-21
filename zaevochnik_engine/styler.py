from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from zaevochnik_engine.lock_rules import resolve_row_lock


def apply_excel_styles(worksheet, start_row: int, end_row: int, column_mapping: dict,
                       excel_styles: dict, validation_prompts: dict,
                       status_rules: list = None, order_type_rules: list = None):
    """Применяет стили оформления, динамически собирая их из внешнего конфигуратора сетки.

    status_rules / order_type_rules - правила полной блокировки + окраски строк
    (см. config.LOCKED_STATUS_RULES / config.LOCKED_ORDER_TYPE_RULES). Если не
    переданы явно - берутся настройки по умолчанию из config.py.
    """
    if status_rules is None or order_type_rules is None:
        from zaevochnik_engine.config import LOCKED_STATUS_RULES, LOCKED_ORDER_TYPE_RULES
        status_rules = status_rules if status_rules is not None else LOCKED_STATUS_RULES
        order_type_rules = order_type_rules if order_type_rules is not None else LOCKED_ORDER_TYPE_RULES

    colors = excel_styles["colors"]
    fonts_cfg = excel_styles["fonts"]
    align_cfg = excel_styles["alignments"]
    heights = excel_styles["row_heights"]

    styles = {
        "header_fill": PatternFill(start_color=colors["primary"], end_color=colors["primary"], fill_type="solid"),
        "header_font": Font(**fonts_cfg["header"]),

        "pack_fill": PatternFill(start_color=colors["bg_pack"], fill_type="solid"),
        "status_new_fill": PatternFill(start_color=colors["bg_new"], fill_type="solid"),
        "default_fill": PatternFill(fill_type=None),

        "regular_font": Font(**fonts_cfg["regular"]),
        "bold_font": Font(**fonts_cfg["bold"]),

        "align_center": Alignment(**align_cfg["default"]),
        "align_left": Alignment(**align_cfg["text_left"]),
        "align_header": Alignment(**align_cfg["header"]),

        "border_data": Border(
            left=Side(border_style="thin", color=colors["border"]),
            right=Side(border_style="thin", color=colors["border"]),
            top=Side(border_style="thin", color=colors["border"]),
            bottom=Side(border_style="thin", color=colors["border"])
        ),
        "border_header": Border(
            left=Side(border_style="thin", color=colors["border"]),
            right=Side(border_style="thin", color=colors["border"]),
            top=Side(border_style="thin", color=colors["border"]),
            bottom=Side(border_style="medium", color=colors["primary"])
        )
    }

    headers = {str(worksheet.cell(row=start_row, column=i).value).strip(): i for i in
               range(1, worksheet.max_column + 1)}
    type_idx = headers.get(column_mapping["type_col"])
    qty_idx = headers.get(column_mapping["qty_col"])
    status_idx = headers.get("Статус")

    # Настройка базовых правил Data Validation
    dv_unit = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
    dv_pack = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")

    for key, dv in [("unit", dv_unit), ("pack", dv_pack)]:
        p = validation_prompts.get(key, {"title": "Внимание", "message": "Заполните поле"})
        dv.promptTitle = p["title"]
        dv.prompt = p["message"]
        dv.errorTitle = "Ошибка ввода"
        dv.error = "Введенное значение не соответствует правилам."
        dv.showInputMessage = True
        dv.showErrorMessage = True
        worksheet.add_data_validation(dv)

    # Оформляем шапку таблицы (строка 5)
    worksheet.row_dimensions[start_row].height = heights["header"]
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=start_row, column=col_idx)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border_header"]
        cell.alignment = styles["align_header"]

    # Итерируемся по строкам данных
    for row_idx in range(start_row + 1, end_row + 1):
        worksheet.row_dimensions[row_idx].height = heights["data"]

        is_pack, is_unit, is_new = False, False, False

        # Считываем статус напрямую из ячейки текущей строки таблицы
        status_text = ""
        if status_idx:
            status_text = str(worksheet.cell(row=row_idx, column=status_idx).value or "").strip().lower()

        # Читаем тип заказа
        type_text = ""
        if type_idx:
            cell_type = worksheet.cell(row=row_idx, column=type_idx)
            current_type_val = str(cell_type.value or "").strip()
            type_text = current_type_val.lower()

            is_pack = "упаковк" in type_text
            is_unit = "штук" in type_text

            if (is_pack or is_unit) and "➔" not in current_type_val:
                cell_type.value = f"{current_type_val} ➔"

        # п.1 + п.2: единая проверка полной блокировки строки по статусу/типу заказа
        is_locked_row, lock_color = resolve_row_lock(
            status_text=status_text,
            type_text=type_text,
            status_rules=status_rules,
            order_type_rules=order_type_rules
        )

        if not is_locked_row and "новинк" in status_text:
            is_new = True

        # Определяем основной цвет фона строки
        if is_locked_row:
            current_fill = PatternFill(start_color=lock_color, end_color=lock_color, fill_type="solid")
        elif is_new:
            current_fill = styles["status_new_fill"]
        elif is_pack or is_unit:
            current_fill = styles["pack_fill"]
        else:
            current_fill = styles["default_fill"]

        # Конфигурируем ячейку ввода количества заказа
        if qty_idx:
            cell_qty = worksheet.cell(row=row_idx, column=qty_idx)

            if is_locked_row:
                cell_qty.value = 0  # Требование 4: правила DataValidation не применяются, значение жестко фиксируется
            elif is_pack:
                dv_pack.add(cell_qty)
                cell_qty.value = None
                # Требование 5: Защита дефолтного стиля от Ctrl+V через условное форматирование
                style_rule = FormulaRule(formula=['1=1'], fill=styles["default_fill"], font=styles["regular_font"],
                                         border=styles["border_data"])
                worksheet.conditional_formatting.add(cell_qty.coordinate, style_rule)
            elif is_unit:
                dv_unit.add(cell_qty)
                cell_qty.value = None
                style_rule = FormulaRule(formula=['1=1'], fill=styles["default_fill"], font=styles["regular_font"],
                                         border=styles["border_data"])
                worksheet.conditional_formatting.add(cell_qty.coordinate, style_rule)

        # Оформление и распределение защиты по всем колонкам строки
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            header_name = str(worksheet.cell(row=start_row, column=col_idx).value or "").strip()
            header_val_lower = header_name.lower()

            # Требование 5: Убираем заливку для колонки количества заказа, если ячейка не заблокирована
            if col_idx == qty_idx and not is_locked_row:
                cell.fill = styles["default_fill"]
            elif current_fill.fill_type:
                cell.fill = current_fill

            cell.font = styles["bold_font"] if "итого" in header_val_lower else styles["regular_font"]
            cell.border = styles["border_data"]

            # Полная сквозная блокировка всей строки при условиях п.1/п.2
            if is_locked_row:
                cell.protection = Protection(locked=True)
            else:
                # В обычном режиме редактировать можно только ячейку количества
                if col_idx == qty_idx:
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)

            if header_name == "Наименование":
                cell.alignment = styles["align_left"]
            else:
                cell.alignment = styles["align_center"]

    # Автоподбор ширины колонок по данным таблицы
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, end_row + 1):
            cell_val = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and not str(cell_val).startswith("="):
                max_len = max(max_len, len(str(cell_val)))

        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
