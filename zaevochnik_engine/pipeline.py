import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from zaevochnik_engine.loader import load_and_clean_data
from zaevochnik_engine.formulas import apply_dynamic_formulas
from zaevochnik_engine.styler import apply_excel_styles
from zaevochnik_engine.cleaner import remove_temporary_columns, remove_rows_by_status
from zaevochnik_engine.header_styler import apply_top_header_and_protection


def build_zaevochnik(
        source_path: str,
        output_path: str,
        sheet_name: str,
        header_row: int,
        filter_col: str,
        filter_val: str,
        drop_columns_finally: list,
        column_mapping: dict,
        weight_column_name: str,
        excel_styles: dict,
        validation_prompts: dict,
        exclude_statuses: list = None
):
    """Изолированный бизнес-процесс сборки с физическим удалением колонок."""
    print("Шаг 1: Загрузка, фильтрация данных и создание структуры...")
    df_clean = load_and_clean_data(
        file_path=source_path,
        filter_column=filter_col,
        filter_value=filter_val,
        qty_col_name=column_mapping["qty_col"],
        total_col_name=column_mapping["total_col"],
        weight_col_name=weight_column_name
    )

    print("Шаг 2: Запись в Excel и ФИЗИЧЕСКОЕ УДАЛЕНИЕ временных столбцов...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=True), start=header_row):
            for c_idx, value in enumerate(row, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        total_rows = header_row + len(df_clean)

        if exclude_statuses:
            total_rows = remove_rows_by_status(
                worksheet=worksheet,
                start_row=header_row,
                total_rows=total_rows,
                status_values=exclude_statuses
            )

        remove_temporary_columns(
            worksheet=worksheet,
            start_row=header_row,
            drop_columns=drop_columns_finally
        )

        print("Шаг 3: Просчет и запись формул на очищенную структуру таблиц...")
        start_data_row = header_row + 1
        apply_dynamic_formulas(
            worksheet=worksheet,
            start_row=start_data_row,
            end_row=total_rows,
            header_row=header_row,
            column_mapping=column_mapping,
            weight_column_name=weight_column_name
        )

        print("Шаг 4: Наложение стилей, подсказок и валидации...")
        apply_excel_styles(
            worksheet=worksheet,
            start_row=header_row,
            end_row=total_rows,
            column_mapping=column_mapping,
            excel_styles=excel_styles,
            validation_prompts=validation_prompts
        )

        print("Шаг 5: Стилизация брендированной шапки макета и включение крипто-защиты листа...")
        apply_top_header_and_protection(
            worksheet=worksheet,
            start_row=header_row,
            end_row=total_rows,
            column_mapping=column_mapping,
            weight_column_name=weight_column_name,
            excel_styles=excel_styles
        )

        from openpyxl.utils import get_column_letter
        max_col_letter = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A{header_row}:{max_col_letter}{total_rows}"

        worksheet.freeze_panes = f"A{start_data_row}"

    print(f"🎉 Процесс завершен! Заявочник обновлен и сохранен: {output_path}")
