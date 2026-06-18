from openpyxl.utils import get_column_letter

def get_excel_col_letter(worksheet, start_row: int, column_name: str) -> str:
    """Определяет текущую букву столбца в Excel по его названию в шапке."""
    for col_idx in range(1, worksheet.max_column + 1):
        if str(worksheet.cell(row=start_row, column=col_idx).value).strip() == column_name:
            return get_column_letter(col_idx)
    raise ValueError(f"Столбец '{column_name}' не найден в итоговой структуре Excel.")


def apply_dynamic_formulas(worksheet, start_row: int, end_row: int, header_row: int, column_mapping: dict,
                           weight_column_name: str):
    """Записывает формулы, защищенные от вырезания (Ctrl+X) с помощью функции ДВССЫЛ (INDIRECT)."""
    f_type = get_excel_col_letter(worksheet, header_row, column_mapping["type_col"])
    f_qty = get_excel_col_letter(worksheet, header_row, column_mapping["qty_col"])
    f_mult = get_excel_col_letter(worksheet, header_row, column_mapping["mult_col"])
    f_total = get_excel_col_letter(worksheet, header_row, column_mapping["total_col"])
    f_weight = get_excel_col_letter(worksheet, header_row, column_mapping["weight_col"])
    f_total_weight = get_excel_col_letter(worksheet, header_row, weight_column_name)

    for i in range(start_row, end_row + 1):
        # Вместо прямой ссылки f"{f_qty}{i}" мы используем INDIRECT(CONCATENATE("{f_qty}", ROW()))
        # В русской версии Excel это автоматически превратится в ДВССЫЛ(СЦЕПИТЬ("буква";СТРОКА()))

        indirect_qty = f'INDIRECT(CONCATENATE("{f_qty}", ROW()))'
        indirect_mult = f'INDIRECT(CONCATENATE("{f_mult}", ROW()))'
        indirect_type = f'INDIRECT(CONCATENATE("{f_type}", ROW()))'
        indirect_weight = f'INDIRECT(CONCATENATE("{f_weight}", ROW()))'
        indirect_total = f'INDIRECT(CONCATENATE("{f_total}", ROW()))'

        # Новая формула для Итого (шт)
        formula_qty = (
            f'=IF(ISNUMBER(SEARCH("упаковка", {indirect_type})), '
            f'IF({indirect_qty}="", 0, {indirect_qty} * {indirect_mult}), '
            f'IF(ISNUMBER(SEARCH("штук", {indirect_type})), '
            f'IF({indirect_qty}="", 0, {indirect_qty}), 0))'
        )

        # Новая формула для Вес (кг)
        formula_weight = (
            f'=IF(OR(ISNUMBER(SEARCH("прямая", {indirect_type})), '
            f'ISNUMBER(SEARCH("приостановлена", {indirect_type}))), 0, '
            f'{indirect_total} * {indirect_weight})'
        )

        # Записываем формулы в ячейки
        worksheet[f"{f_total}{i}"] = formula_qty
        worksheet[f"{f_total_weight}{i}"] = formula_weight